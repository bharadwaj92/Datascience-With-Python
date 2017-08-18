# importing library files
import pandas as pd
import requests
import os
import zipfile
import openpyxl
import sqlite3
import glob
import getpass
import xlrd
import numpy as np
#library for converting dataframe to rows for loading to excel workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import gc
# global declaration section
# contains links and locations
download_url = url="https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"
k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"
staging_dir_name = "staging"
focussed_states = {}
#Method that helps download the files from web.
# Extracts the files and puts in staging directory
def download_webfiles():
    os.mkdir(staging_dir_name)
    zip_file_name = os.path.join(staging_dir_name, "test.zip")
    r = requests.get(url)
    zf = open(zip_file_name, "wb")
    zf.write(r.content)
    zf.close()
    z = zipfile.ZipFile(zip_file_name, 'r')
    z.extractall(staging_dir_name)
    z.close()
    xf = open("hospital_ranking_focus_states.xlsx", "wb")
    r = requests.get(k_url)
    xf.write(r.content)
    xf.close()
    ##Enables garbage collection
    gc.collect()
    return

## Method that helps creating measures_statistics.
## Contains two queries that are needed for population nationwide sheet
## as well as contains query for populating statistical measures for each state
def create_second_sheet():
    res = []
    query1 = "select t.measure_id  as 'Measure ID', t.measure_name  as 'Measure Name'," + \
    "min(t.score) as 'Minimum' , max(t.score) as 'Maximum', avg(t.score) as 'Average'," + \
    "AVG((t.score - s.a) * (t.score - s.a)) as 'vari'" + \
    "from timely_and_effective_care___hospital t ," + \
    "(select measure_id as 'm' , avg(score) as 'a' from timely_and_effective_care___hospital group by measure_id) as s " + \
    "WHERE t.measure_id = s.m and " + \
    "CAST(score AS float) IS score group by 1,2;"
    query2 = "select t.measure_id  as 'Measure ID', t.measure_name  as 'Measure Name',"+\
        "min(t.score) as 'Minimum' , max(t.score) as 'Maximum', avg(t.score) as 'Average',"+\
        "AVG((t.score - s.a) * (t.score - s.a)) as 'vari'"+\
        "from timely_and_effective_care___hospital t ,"+\
        "(select measure_id as 'm' , avg(score) as 'a' from timely_and_effective_care___hospital group by measure_id) as s "+\
        "WHERE t.measure_id = s.m and t.state = ? and" +\
        " CAST(score AS float) IS score group by 1,2;"
    conn = sqlite3.connect("medicare_hospital_compare.db")
    headers = ["Measure ID", "Measure Name", "Minimum", "Maximum", "Average", "Standard Deviation"]
    outputsheet = openpyxl.Workbook()
    sheet_1 = outputsheet.create_sheet("Nationwide")
    rows = conn.execute(query1)
    # appends rows to result
    for row in rows:
        res.append(row)
    ## Calculates the standard deviation from variance
    ## Converts the string variance to float
    ## then calculates sqrt of var using numpy library
    ## Converts back to string for writing to excel sheet
    outputdf1 = pd.DataFrame(res,index=None,columns=headers, dtype='str')
    outputdf1['Standard Deviation']= outputdf1['Standard Deviation'].astype(float).fillna(0.0)
    outputdf1['Standard Deviation'] = np.sqrt(outputdf1['Standard Deviation'])
    outputdf1['Standard Deviation'] = outputdf1['Standard Deviation'].astype('str')
    for r in dataframe_to_rows(outputdf1, index=False, header=True):
        sheet_1.append(r)
    ## loops over each state and creates a sheet for each state with statistical measures
    for state , stabb in sorted(focussed_states.items()):
        result_state = []
        sheet_state = outputsheet.create_sheet(state)
        rows = conn.execute(query2,(stabb.strip(),))
        for row in rows:
            result_state.append(row)
        ## Calculates the standard deviation from variance
        ## Converts the string variance to float
        ## then calculates sqrt of var using numpy library
        ## Converts back to string for writing to excel sheet
        outputdf2 = pd.DataFrame(result_state, index=None, columns=headers, dtype='str')
        outputdf2['Standard Deviation']= outputdf2['Standard Deviation'].astype(float).fillna(0.0)
        outputdf2['Standard Deviation'] = np.sqrt(outputdf2['Standard Deviation'])
        outputdf2['Standard Deviation'] = outputdf2['Standard Deviation'].astype('str')
        for r in dataframe_to_rows(outputdf2, index=False, header=True):
            sheet_state.append(r)
    ## Save the sheet
    outputsheet.save("measures_statistics.xlsx")
    wb = openpyxl.load_workbook("measures_statistics.xlsx")
    rm = wb.get_sheet_by_name("Sheet")
    wb.remove_sheet(rm)
    print(wb.get_sheet_names())
    wb.save("measures_statistics.xlsx")
    ## close the database connection
    conn.close()
    ##Enables garbage collection
    gc.collect()
    return

## Method that creates first sheet hospital_ranking.
## Contains queries and creating workbook stuff for providers as well as focus states
def create_output_sheet():
    headers_1 = ["Provider ID","Hospital Name", "City", "State","County"]
    wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")
    input_sheet1 = wb.get_sheet_by_name("Hospital National Ranking")
    input_sheet2 = wb.get_sheet_by_name("Focus States")
    query1 = "select provider_id as 'Provider ID', hospital_name as 'Hospital Name'"+\
            ",city as 'City', state as 'State', county_name as 'County' from hospital_general_information where"+ \
            " trim(provider_id) ="
    query2 = "select provider_id as 'Provider ID', hospital_name as 'Hospital Name'" + \
             ",city as 'City', state as 'State', county_name as 'County' from hospital_general_information" +\
              " where state ="
    i =2
    result = []
    conn = sqlite3.connect("medicare_hospital_compare.db")
    while (i<=101):
        query = query1 + "'"+(input_sheet1.cell(row=i, column=1).value).strip() + "'"
        rows = conn.execute(query)
        for row in rows:
            result.append(row)
        i += 1
    outputdf = pd.DataFrame(result,columns=headers_1,dtype='str')
    outputsheet =openpyxl.Workbook()
    sheet_1 = outputsheet.create_sheet("Nationwide")
    for r in dataframe_to_rows(outputdf, index=False, header=True):
        sheet_1.append(r)
    j=2
    while input_sheet2.cell(row=j, column=1).value != None:
        focussed_states[input_sheet2.cell(row=j, column=1).value.strip()] = input_sheet2.cell(row=j, column=2).value.strip()
        j += 1
    ## Foreach item in sorted focus states, check for provider id using loop
    ## loop through the result set and check if count of 100 is obtained
    ## Continue if not else break and goto next state
    for k , v in sorted(focussed_states.items()):
        print("running for ",k)
        sheet_count = 2
        row_count = 1
        res = []
        while (row_count <= 100):
            query= query2
            query = query + "'"+ v + "' and trim(provider_id)=" + "'" + (input_sheet1.cell(row=sheet_count, column=1).value).strip() + "'"
            rows = conn.execute(query)
            x = [x for x in rows]
            if(len(x) ==0):
                sheet_count += 1
                continue
            else:
                res.append(x[0])
                sheet_count += 1
                row_count += 1
        outputdf_state = pd.DataFrame(res,columns=headers_1,dtype='str')
        sheet_2 = outputsheet.create_sheet(k)
        for r in dataframe_to_rows(outputdf_state, index=False, header=True):
            sheet_2.append(r)
    ## save the sheet and close the database connection
    outputsheet.save("hospital_ranking.xlsx")
    wb = openpyxl.load_workbook("hospital_ranking.xlsx")
    rm = wb.get_sheet_by_name("Sheet")
    wb.remove_sheet(rm)
    print(wb.get_sheet_names())
    wb.save("hospital_ranking.xlsx")
    conn.close()
    ##Enables garbage collection
    gc.collect()
    return

## Function that creates tables from the csv files
## Checks if the file is cp1252 encoded and converts that to UTF-8 and then loads to database
## Raises system exit if the database connection could not be obtained.
def create_tables():
    try:
        conn = sqlite3.connect("medicare_hospital_compare.db")
        glob_dir = os.path.join(staging_dir_name, "*.csv")
        for file_name in glob.glob(glob_dir):
            absolute_path = os.path.abspath(file_name)
            try:
                df= pd.read_csv(absolute_path)
                colnames = list(df.columns)
                dict_types = {x:'str' for x in colnames}
                df = pd.read_csv(absolute_path,dtype=dict_types)
            except UnicodeDecodeError:
                in_fp = open(absolute_path, "rt", encoding='cp1252')  # rt = read text
                input_data = in_fp.read()
                in_fp.close()
                newfilename = absolute_path + ".fix"
                out_fp = open(newfilename, "wt", encoding='utf-8')
                for c in input_data:
                    if c != '\0':
                        out_fp.write(c)
                out_fp.close()
                df = pd.read_csv(newfilename)
                colnames = list(df.columns)
                dict_types = {x:'str' for x in colnames}
                df = pd.read_csv(newfilename,dtype=dict_types)
            colnames = list(df.columns)
            col_names_new = []
            for item in colnames:
                temp = item.lower().replace(" ", "_").replace("-", "_").replace("%", "pct").replace("/", "_")
                if(not temp[0].isalpha()):
                    col_names_new.append("c_" + temp)
                else:
                    col_names_new.append(temp)
            df.columns = col_names_new
            tablename = os.path.splitext(os.path.basename(file_name))[0].lower().replace(" ","_").replace("-","_").replace("%", "pct").replace("/","_")
            if(not tablename[0].isalpha()):
                tablename = "t_" + tablename
            df.to_sql(tablename,con=conn,if_exists='fail',dtype = {col:'text' for col in df}, index=False)
    ## Raises exception on unable to establish connection
    except ConnectionError:
        print("Unable to get database connection.. Exiting out of program")
        raise SystemExit
    ## Closes connection upon running the query
    finally:
        conn.close()
    ##Enables garbage collection
    gc.collect()
    return

## Driver function which calls other functions.
if __name__ == "__main__":
    download_webfiles()
    create_tables()
    create_output_sheet()
    create_second_sheet()
    ##Enables garbage collection
    gc.collect()
